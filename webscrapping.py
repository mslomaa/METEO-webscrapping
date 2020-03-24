from bs4 import BeautifulSoup as soup
import requests
import numpy as np

r = requests.get('https://danepubliczne.imgw.pl/api/data/synop/format/html')

page_soup = soup(r.text, 'html.parser')
dane = page_soup.findAll('tr')[1:]

srednia_temp_list = []
srednia_predkosc_wiatru = []
sredni_kierunek_wiatru = []
srednia_wilgotnosc = []
srednie_cisnienie = []
suma_opadu_cal = []

for dan in dane:
    nowy = [i.text for i in dan.find_all('td')]
    id_stacji = nowy[0]
    miasto = nowy[1]
    data_pomiaru = nowy[2]
    godzina = nowy[3]
    temperatura = nowy[4]
    srednia_temp_list.append(temperatura)
    predkosc_wiatru = nowy[5]
    srednia_predkosc_wiatru.append(predkosc_wiatru)
    kierunek_wiatru = nowy[6]
    sredni_kierunek_wiatru.append(kierunek_wiatru)
    wilgotnosc = nowy[7]
    if wilgotnosc.strip():
        srednia_wilgotnosc.append(wilgotnosc)
    suma_opadu = nowy[8]
    if suma_opadu.strip():
        suma_opadu_cal.append(suma_opadu)
    cisnienie = nowy[9]
    if cisnienie.strip():
        srednie_cisnienie.append(cisnienie)

srednia_temp_wyniki = []
suma_temp = 0

for temp in srednia_temp_list:
    pomiar_temp = float(temp)
    suma_temp += pomiar_temp
    srednia = suma_temp / len(srednia_temp_list)
    srednia_temp_wyniki.append(srednia)
print("Średnia temperatura w Polsce o godz: " + str(godzina) + ' to: ' + str(round(srednia_temp_wyniki.pop(), 4)) + " stopni C")


srednia_predkosc_wiatru_wyniki = []
suma_wiatr = 0
for wiatr in srednia_predkosc_wiatru:
    pomiar_wiatr = int(wiatr)
    suma_wiatr += pomiar_wiatr
    srednia = suma_wiatr / len(srednia_predkosc_wiatru)
    srednia_predkosc_wiatru_wyniki.append(srednia)
print("Średnia predkość wiatru w Polsce o godz: " + str(godzina) + ' to: ' + str(round(srednia_predkosc_wiatru_wyniki.pop(), 3)) + " w Skal Beauforta")


sredni_kierunek_wiatru_wyniki = []
suma_kier_wiatr = 0
for kierunek in sredni_kierunek_wiatru:
    pomiar_kierunek = int(kierunek)
    suma_kier_wiatr += pomiar_kierunek
    srednia = suma_kier_wiatr / len(sredni_kierunek_wiatru)
    sredni_kierunek_wiatru_wyniki.append(srednia)
#     # if srednia == range(0, 46) or srednia == range(320, 361):
#     #     srednia = "N"
#     # elif srednia == range(46):
#
print("Średni kierunek wiatru w Polsce o godz: " + str(godzina) + ' to: ' + str(round(sredni_kierunek_wiatru_wyniki.pop(), 3)) + " stopni")


srednia_wilgotnosc_wyniki = []
suma_wilg = 0
for wilgotn in srednia_wilgotnosc:
    # if wilgotn.strip():
        pomiar_wilg = float(wilgotn)
        suma_wilg += pomiar_wilg
        srednia = suma_wilg / len(srednia_wilgotnosc)
        srednia_wilgotnosc_wyniki.append(srednia)
print("Średnia wilgotność powietrza w Polsce o godz: " + str(godzina) + ' to: ' + str(round(srednia_wilgotnosc_wyniki.pop(), 2)) + " %")


srednie_cisnienie_wyniki = []
suma_cis = 0
for cis in srednie_cisnienie:
        pomiar_cis = float(cis)
        suma_cis += pomiar_cis
        srednia = suma_cis / len(srednie_cisnienie)
        srednie_cisnienie_wyniki.append(srednia)
print("Średnie ciśnienie w Polsce o godz: " + str(godzina) + ' to: ' + str(round(srednie_cisnienie_wyniki.pop(), 2)) + " hPa")


suma_opad = 0
for opad in suma_opadu_cal:
    suma_opad += float(opad)
print('Całkowita suma opadu w Polsce do godz: ' + str(godzina) + ' to: ' + str(suma_opad) + ' mm')

from openpyxl import Workbook

dane_excel = Workbook()
dane_excel_zakladka = dane_excel.active

dane_update = (
    ['Średnia temperatura', str(data_pomiaru), str(godzina), str(srednia_temp_wyniki.pop())],
    ['Średnia prędkość wiatru', str(data_pomiaru), str(godzina), str(srednia_predkosc_wiatru_wyniki.pop())],
    ['Średni kierunek wiatru', str(data_pomiaru), str(godzina), str(sredni_kierunek_wiatru_wyniki.pop())],
    ['Średnia wilgotność', str(data_pomiaru), str(godzina), str(srednia_wilgotnosc_wyniki.pop())],
    ['Średnie ciśnienie', str(data_pomiaru), str(godzina), str(srednie_cisnienie_wyniki.pop())],
    ['Całkowita supa opadu', str(data_pomiaru), str(godzina), str(suma_opad)],
)

row = 0
column = 0

for parametr, data, godz, wynik in dane_update:
    dane_excel_zakladka.cell(row=1, column=1).value = parametr
    dane_excel_zakladka.cell(row=2, column=2).value = data
    dane_excel_zakladka.cell(row=3, column=3).value = godz
    dane_excel_zakladka.cell(row=4, column=4).value = wynik
    row += 1

dane_excel.save('meteo.xlsx')